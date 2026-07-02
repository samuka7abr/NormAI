import csv
import io

from openpyxl import Workbook, load_workbook

from infrastructure.processor.normalization_processor import (
    CSV_CONTENT_TYPE,
    XLSX_CONTENT_TYPE,
    NormalizationProcessor,
)


def _read_csv(content: bytes) -> list[dict[str, str]]:
    return list(csv.DictReader(io.StringIO(content.decode("utf-8"))))


async def test_normalization_processor_applies_csv_rules():
    processor = NormalizationProcessor()
    content = (
        "tribunal,comarca,especies_afetadas,nulo,raw\n"
        "\"Tribunal de Justiça do Estado de São Paulo TJSP\","
        "\"  poços   de caldas  \","
        "\"Cachorro|galgo inglês\","
        "\"Não informado\","
        "\"  manter espaços  \"\n"
    ).encode("utf-8")

    result = await processor.process(
        content=content,
        original_filename="relatorio.csv",
        column_config_snapshot={
            "tribunal": {
                "enabled": True,
                "normalizations": {"trim": True, "abbreviate": True},
            },
            "comarca": {
                "enabled": True,
                "normalizations": {"trim": True, "capitalize_pt_br": True},
            },
            "especies_afetadas": {
                "enabled": True,
                "normalizations": {"trim": True, "split": True, "capitalize_pt_br": True},
            },
            "nulo": {
                "enabled": True,
                "normalizations": {"trim": True, "nulls": True},
            },
            "raw": {
                "enabled": False,
                "normalizations": {"trim": True},
            },
        },
    )

    rows = _read_csv(result.content)
    assert result.filename == "normalized_relatorio.csv"
    assert result.content_type == CSV_CONTENT_TYPE
    assert rows[0]["tribunal"] == "TJSP"
    assert rows[0]["comarca"] == "Poços de Caldas"
    assert rows[0]["especies_afetadas"] == "Cachorro|Galgo Inglês"
    assert rows[0]["nulo"] == ""
    assert rows[0]["raw"] == "  manter espaços  "


async def test_normalization_processor_applies_suffix_and_accent_rules():
    processor = NormalizationProcessor()
    content = (
        "tribunal,classe\n"
        "\"TJMG Comarca de Itamogi\",\"  Ação   Cível  \"\n"
        "\"TRF 1ª Região / Seção Judiciária\",\"Agravo de Instrumento\"\n"
    ).encode("utf-8")

    result = await processor.process(
        content=content,
        original_filename="relatorio.csv",
        column_config_snapshot={
            "tribunal": {
                "enabled": True,
                "normalizations": {
                    "trim": True,
                    "remove_suffixes": True,
                    "abbreviate": True,
                },
            },
            "classe": {
                "enabled": True,
                "normalizations": {
                    "trim": True,
                    "remove_accents": True,
                    "capitalize_pt_br": True,
                },
            },
        },
    )

    rows = _read_csv(result.content)
    assert rows[0]["tribunal"] == "TJMG"
    assert rows[0]["classe"] == "Acao Civel"
    assert rows[1]["tribunal"] == "TRF1"
    assert rows[1]["classe"] == "Agravo de Instrumento"


async def test_normalization_processor_respects_custom_split_separators():
    processor = NormalizationProcessor()
    content = (
        "especies_afetadas\n"
        "\"Cachorro|Gato;Ave (sapo/rá)|Não informado\"\n"
    ).encode("utf-8")

    result = await processor.process(
        content=content,
        original_filename="relatorio.csv",
        column_config_snapshot={
            "especies_afetadas": {
                "enabled": True,
                "normalizations": {
                    "trim": True,
                    "nulls": True,
                    "split": True,
                    "split_separators": ["|", ";"],
                },
            },
        },
    )

    rows = _read_csv(result.content)
    assert rows[0]["especies_afetadas"] == "Cachorro|Gato|Ave (sapo/rá)"


async def test_normalization_processor_applies_xlsx_rules():
    processor = NormalizationProcessor()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["comarca", "valor"])
    sheet.append(["ribeirão preto", "NA"])

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    result = await processor.process(
        content=stream.getvalue(),
        original_filename="relatorio.xlsx",
        column_config_snapshot={
            "comarca": {
                "enabled": True,
                "normalizations": {"capitalize_pt_br": True},
            },
            "valor": {
                "enabled": True,
                "normalizations": {"nulls": True},
            },
        },
    )

    output = load_workbook(io.BytesIO(result.content))
    sheet = output.active

    assert result.filename == "normalized_relatorio.xlsx"
    assert result.content_type == XLSX_CONTENT_TYPE
    assert sheet["A2"].value == "Ribeirão Preto"
    assert sheet["B2"].value is None

    output.close()
