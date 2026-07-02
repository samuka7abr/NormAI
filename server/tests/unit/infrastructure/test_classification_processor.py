import csv
import io
from collections.abc import Callable

from openpyxl import Workbook, load_workbook

from application.classification.classify_column import ClassifyColumnValuesUseCase
from application.classification.schemas import (
    CategoryDiscoveryResponse,
    ClassifiedItem,
    ClassifyBatchResponse,
)
from domain.shared.llm_client import LLMClient, LLMError, LLMMessage
from infrastructure.processor.classification_processor import ClassificationProcessor
from infrastructure.processor.normalization_processor import NormalizationProcessor


class FakeLLMClient(LLMClient):
    def __init__(
        self,
        *,
        discovery: Callable[[list[LLMMessage]], CategoryDiscoveryResponse] | None = None,
        classify: Callable[[list[LLMMessage]], ClassifyBatchResponse] | None = None,
    ) -> None:
        self.discovery_handler = discovery
        self.classify_handler = classify
        self.system_prompts: list[str] = []

    async def chat_structured(self, messages, schema, *, temperature: float = 0.0):
        self.system_prompts.append(next(m.content for m in messages if m.role == "system"))
        if schema is CategoryDiscoveryResponse:
            assert self.discovery_handler is not None
            return self.discovery_handler(messages)
        if schema is ClassifyBatchResponse:
            assert self.classify_handler is not None
            return self.classify_handler(messages)
        raise AssertionError(f"unexpected schema {schema}")


def _extract_batch_values(messages: list[LLMMessage]) -> list[str]:
    user = next(m for m in messages if m.role == "user")
    return [
        line.removeprefix("- ").strip()
        for line in user.content.splitlines()
        if line.startswith("- ")
    ]


def _build_processor(
    *,
    categories: list[str],
    mapping: dict[str, str],
) -> tuple[ClassificationProcessor, FakeLLMClient]:
    def discovery(_messages):
        return CategoryDiscoveryResponse(categories=categories)

    def classify(messages):
        batch = _extract_batch_values(messages)
        return ClassifyBatchResponse(
            items=[ClassifiedItem(value=v, category=mapping[v]) for v in batch if v in mapping]
        )

    llm = FakeLLMClient(discovery=discovery, classify=classify)
    use_case = ClassifyColumnValuesUseCase(llm, sample_size=50, batch_size=20)
    processor = ClassificationProcessor(
        normalization_processor=NormalizationProcessor(),
        classify_use_case=use_case,
    )
    return processor, llm


def _read_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
    return list(reader.fieldnames or []), list(reader)


async def test_csv_pipeline_normalizes_and_classifies():
    processor, llm = _build_processor(
        categories=["Cães", "Aves"],
        mapping={
            "Pitbull": "Cães",
            "Galgo Inglês": "Cães",
            "Pombo": "Aves",
        },
    )

    content = (
        "tribunal,especies\n"
        "Tribunal de Justiça do Estado de São Paulo TJSP,Pitbull\n"
        "TJSP,Pitbull\n"
        "tribunal de justiça do estado de são paulo,Galgo Inglês\n"
        "TJSP,Pombo\n"
    ).encode("utf-8")

    snapshot = {
        "tribunal": {
            "enabled": True,
            "normalizations": {"trim": True, "abbreviate": True},
            "classify": False,
            "categories": None,
        },
        "especies": {
            "enabled": True,
            "normalizations": {"trim": True},
            "classify": True,
            "categories": None,
        },
        "_project": {"ai_context": "Foco em maus tratos a animais."},
    }

    result = await processor.process(
        content=content,
        original_filename="rel.csv",
        column_config_snapshot=snapshot,
    )

    headers, rows = _read_csv(result.content)
    assert headers == ["tribunal", "especies", "especies_categoria"]
    assert [r["tribunal"] for r in rows] == ["TJSP", "TJSP", "TJSP", "TJSP"]
    assert [r["especies"] for r in rows] == ["Pitbull", "Pitbull", "Galgo Inglês", "Pombo"]
    assert [r["especies_categoria"] for r in rows] == ["Cães", "Cães", "Cães", "Aves"]

    # ai_context entrou no system prompt
    assert any("maus tratos a animais" in p for p in llm.system_prompts)
    # last_metrics expõe o ClassificationResult pra persistência
    assert "especies" in processor.last_metrics
    assert "Cães" in processor.last_metrics["especies"].categories


async def test_pipeline_skips_classification_when_no_targets():
    processor, llm = _build_processor(
        categories=["Cães"],
        mapping={"Pitbull": "Cães"},
    )

    content = b"tribunal\nTJSP\n"
    snapshot = {
        "tribunal": {
            "enabled": True,
            "normalizations": {"trim": True},
            "classify": False,
            "categories": None,
        }
    }

    result = await processor.process(
        content=content,
        original_filename="rel.csv",
        column_config_snapshot=snapshot,
    )

    headers, _rows = _read_csv(result.content)
    assert headers == ["tribunal"]
    assert llm.system_prompts == []
    assert processor.last_metrics == {}


async def test_empty_values_dont_classify():
    """Linhas com valor vazio na coluna classify ficam com categoria vazia."""
    processor, _llm = _build_processor(
        categories=["Cães"],
        mapping={"Pitbull": "Cães"},
    )

    content = (
        "nome,especies\n"
        "a,Pitbull\n"
        "b,\n"
        "c,Pitbull\n"
    ).encode("utf-8")
    snapshot = {
        "especies": {
            "enabled": True,
            "normalizations": {},
            "classify": True,
            "categories": None,
        }
    }

    result = await processor.process(
        content=content,
        original_filename="rel.csv",
        column_config_snapshot=snapshot,
    )

    _headers, rows = _read_csv(result.content)
    assert [r["especies_categoria"] for r in rows] == ["Cães", "", "Cães"]


async def test_xlsx_pipeline_adds_category_column():
    processor, _llm = _build_processor(
        categories=["Aves"],
        mapping={"Pombo": "Aves", "Papagaio": "Aves"},
    )

    wb = Workbook()
    ws = wb.active
    ws.append(["especies"])
    ws.append(["Pombo"])
    ws.append(["Papagaio"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()

    snapshot = {
        "especies": {
            "enabled": True,
            "normalizations": {},
            "classify": True,
            "categories": None,
        }
    }
    result = await processor.process(
        content=buf.getvalue(),
        original_filename="rel.xlsx",
        column_config_snapshot=snapshot,
    )

    workbook = load_workbook(io.BytesIO(result.content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    body = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    workbook.close()

    assert headers == ["especies", "especies_categoria"]
    assert body == [["Pombo", "Aves"], ["Papagaio", "Aves"]]


async def test_llm_failure_falls_back_to_others_without_crashing():
    def discovery(_messages):
        raise LLMError("simulated outage")

    llm = FakeLLMClient(discovery=discovery)
    use_case = ClassifyColumnValuesUseCase(llm, sample_size=10, batch_size=5)
    processor = ClassificationProcessor(
        normalization_processor=NormalizationProcessor(),
        classify_use_case=use_case,
    )

    content = b"especies\nPitbull\nPombo\n"
    snapshot = {
        "especies": {
            "enabled": True,
            "normalizations": {},
            "classify": True,
            "categories": None,
        }
    }
    result = await processor.process(
        content=content,
        original_filename="rel.csv",
        column_config_snapshot=snapshot,
    )

    _headers, rows = _read_csv(result.content)
    # Sem LLM, todos viram "Outros" e a execução não crasha.
    assert [r["especies_categoria"] for r in rows] == ["Outros", "Outros"]
    assert processor.last_metrics["especies"].failed_values  # marcados como falha


async def test_classify_on_missing_column_is_skipped():
    processor, llm = _build_processor(
        categories=["Cães"],
        mapping={"Pitbull": "Cães"},
    )

    content = b"tribunal,especies\nTJSP,Pitbull\n"
    snapshot = {
        "outra_coluna": {  # coluna que não existe no arquivo
            "enabled": True,
            "normalizations": {},
            "classify": True,
            "categories": None,
        }
    }
    result = await processor.process(
        content=content,
        original_filename="rel.csv",
        column_config_snapshot=snapshot,
    )

    headers, _rows = _read_csv(result.content)
    # Não adicionou coluna nova e não chamou LLM
    assert "outra_coluna_categoria" not in headers
    assert llm.system_prompts == []
