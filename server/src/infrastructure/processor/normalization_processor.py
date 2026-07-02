import csv
import io
import re
import unicodedata
from collections.abc import Iterable
from typing import Any

from application.reports.report_processor import ProcessingResult, ReportProcessor


CSV_CONTENT_TYPE = "text/csv; charset=utf-8"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

DEFAULT_SPLIT_SEPARATORS = ("|", ";", "/")

NULL_ALIASES = {
    "",
    "-",
    "--",
    "n/a",
    "na",
    "nan",
    "null",
    "none",
    "nao informado",
    "nao informada",
    "nao informa",
    "nao se aplica",
    "nao aplicavel",
    "nao especificado",
    "nao especificada",
    "nenhum",
    "nenhuma",
    "sem informacao",
    "sem informação",
    "sem dados",
}

LOWERCASE_WORDS = {
    "a",
    "as",
    "com",
    "da",
    "das",
    "de",
    "do",
    "dos",
    "e",
    "em",
    "na",
    "nas",
    "no",
    "nos",
    "o",
    "os",
    "para",
    "por",
}

STATE_TO_UF = {
    "acre": "AC",
    "alagoas": "AL",
    "amapa": "AP",
    "amazonas": "AM",
    "bahia": "BA",
    "ceara": "CE",
    "distrito federal": "DF",
    "espirito santo": "ES",
    "goias": "GO",
    "maranhao": "MA",
    "mato grosso": "MT",
    "mato grosso do sul": "MS",
    "minas gerais": "MG",
    "para": "PA",
    "paraiba": "PB",
    "parana": "PR",
    "pernambuco": "PE",
    "piaui": "PI",
    "rio de janeiro": "RJ",
    "rio grande do norte": "RN",
    "rio grande do sul": "RS",
    "rondonia": "RO",
    "roraima": "RR",
    "santa catarina": "SC",
    "sao paulo": "SP",
    "sergipe": "SE",
    "tocantins": "TO",
}

KNOWN_ACRONYM_RE = re.compile(
    r"\b("
    r"STF|STJ|TST|TSE|STM|CNJ|"
    r"TRF[1-6]?|TRT\d{1,2}|TRE[- ]?[A-Z]{2}|"
    r"TJ[A-Z]{2}|TJDFT|"
    r"MPF|MPT|MP[A-Z]{2}|DPU|DPE[- ]?[A-Z]{2}"
    r")\b",
    re.IGNORECASE,
)

TRF_REGION_RE = re.compile(
    r"\btribunal\s+regional\s+federal\b.*?\b([1-6])\s*(?:a|ª|o|º)?\s+regi[aã]o\b",
    re.IGNORECASE,
)
TRF_SHORT_REGION_RE = re.compile(
    r"\bTRF\s*([1-6])\s*(?:a|ª|o|º)?\s+regi[aã]o\b",
    re.IGNORECASE,
)


class NormalizationProcessor(ReportProcessor):
    async def process(
        self,
        content: bytes,
        original_filename: str,
        column_config_snapshot: dict,
    ) -> ProcessingResult:
        if original_filename.lower().endswith(".xlsx"):
            return _process_xlsx(content, original_filename, column_config_snapshot)
        return _process_csv(content, original_filename, column_config_snapshot)


def _process_csv(
    content: bytes,
    original_filename: str,
    column_config_snapshot: dict,
) -> ProcessingResult:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        return ProcessingResult(
            content=b"",
            filename=_result_filename(original_filename),
            content_type=CSV_CONTENT_TYPE,
        )

    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=reader.fieldnames, lineterminator="\n")
    writer.writeheader()

    for row in reader:
        normalized = {
            column: _normalize_cell(value, column_config_snapshot.get(column, {}))
            for column, value in row.items()
            if column in reader.fieldnames
        }
        writer.writerow(normalized)

    return ProcessingResult(
        content=output.getvalue().encode("utf-8"),
        filename=_result_filename(original_filename),
        content_type=CSV_CONTENT_TYPE,
    )


def _process_xlsx(
    content: bytes,
    original_filename: str,
    column_config_snapshot: dict,
) -> ProcessingResult:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row):
            if index >= len(headers):
                continue
            header = headers[index]
            if header is None:
                continue
            cell.value = _normalize_cell(cell.value, column_config_snapshot.get(str(header), {}))

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    return ProcessingResult(
        content=output.getvalue(),
        filename=_result_filename(original_filename),
        content_type=XLSX_CONTENT_TYPE,
    )


def _normalize_cell(value: Any, config: dict) -> str:
    if not config or not config.get("enabled", True):
        return "" if value is None else str(value)

    normalizations = config.get("normalizations") or {}
    text = "" if value is None else str(value)

    if _enabled(normalizations, "trim", "trim_whitespace", "strip", "remover_espacos"):
        text = _trim(text)

    if _enabled(normalizations, "nulls", "null_aliases", "treat_nulls", "tratar_nulos"):
        if _is_null_alias(text):
            return ""

    if _enabled(normalizations, "split", "split_separators", "split_by_separator"):
        return _normalize_split_values(text, config, normalizations)

    return _normalize_scalar(text, normalizations)


def _normalize_split_values(text: str, config: dict, normalizations: dict) -> str:
    token_config = {
        **config,
        "normalizations": {
            **normalizations,
            "split": False,
            "split_separators": False,
            "split_by_separator": False,
        },
    }
    tokens = [
        _normalize_cell(token, token_config)
        for token in _split_text(text, _split_separators(normalizations))
    ]
    return "|".join(token for token in tokens if token)


def _normalize_scalar(text: str, normalizations: dict) -> str:
    if _enabled(normalizations, "remove_suffixes", "remove_suffix", "remover_sufixos"):
        text = _remove_suffixes(text)

    if _enabled(normalizations, "abbreviate", "abbreviation", "abreviar", "abbreviacao"):
        text = _abbreviate(text)

    if _enabled(normalizations, "remove_accents", "strip_accents", "remover_acentos"):
        text = _remove_accents(text)

    if _enabled(normalizations, "capitalize_pt_br", "capitalization_pt_br", "capitalizar"):
        text = _capitalize_pt_br(text)

    return text


def _enabled(normalizations: dict, *keys: str) -> bool:
    for key in keys:
        value = normalizations.get(key)
        if isinstance(value, bool):
            return value
        if value is not None:
            return bool(value)
    return False


def _trim(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _is_null_alias(text: str) -> bool:
    normalized = _remove_accents(_trim(text)).lower()
    return normalized in {_remove_accents(alias).lower() for alias in NULL_ALIASES}


def _split_text(text: str, separators: Iterable[str]) -> list[str]:
    escaped = [re.escape(separator) for separator in separators if separator]
    if not escaped:
        return [text]
    return re.split("|".join(escaped), text)


def _split_separators(normalizations: dict) -> tuple[str, ...]:
    value = (
        normalizations.get("separators")
        or normalizations.get("split_separators")
        or normalizations.get("separadores")
    )
    if isinstance(value, str):
        return tuple(value)
    if isinstance(value, list | tuple):
        return tuple(str(item) for item in value)
    return DEFAULT_SPLIT_SEPARATORS


def _remove_suffixes(text: str) -> str:
    stripped = _trim(text)
    trf_short_region = TRF_SHORT_REGION_RE.search(stripped)
    if trf_short_region:
        return f"TRF{trf_short_region.group(1)}"

    acronym = KNOWN_ACRONYM_RE.match(stripped)
    if acronym:
        return acronym.group(1).replace(" ", "").replace("-", "").upper()

    for separator in (" - ", "-", "/"):
        if separator in stripped:
            return _trim(stripped.split(separator, 1)[0])
    return stripped


def _abbreviate(text: str) -> str:
    stripped = _trim(text)

    trf_short_region = TRF_SHORT_REGION_RE.search(stripped)
    if trf_short_region:
        return f"TRF{trf_short_region.group(1)}"

    trf_region = TRF_REGION_RE.search(stripped)
    if trf_region:
        return f"TRF{trf_region.group(1)}"

    acronym = KNOWN_ACRONYM_RE.search(stripped)
    if acronym:
        return acronym.group(1).replace(" ", "").replace("-", "").upper()

    comparable = _remove_accents(stripped).lower()
    if "tribunal de justica" in comparable:
        for state, uf in STATE_TO_UF.items():
            if state in comparable:
                return f"TJ{uf}"

    return stripped


def _remove_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def _capitalize_pt_br(text: str) -> str:
    words = _trim(text).lower().split(" ")
    capitalized: list[str] = []
    for index, word in enumerate(words):
        if not word:
            continue
        if index > 0 and word in LOWERCASE_WORDS:
            capitalized.append(word)
            continue
        capitalized.append(_capitalize_hyphenated(word))
    return " ".join(capitalized)


def _capitalize_hyphenated(word: str) -> str:
    return "-".join(part[:1].upper() + part[1:] if part else part for part in word.split("-"))


def _result_filename(original_filename: str) -> str:
    return f"normalized_{original_filename}"
