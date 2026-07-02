"""Utilidades para ler/escrever CSV e XLSX como tabelas estruturadas.

Modelo intermediário: `(headers: list[str], rows: list[dict[str, str]])`.
Tudo que entra/sai é string — normalização e classificação só trabalham com texto.
"""
import csv
import io
from dataclasses import dataclass

CSV_CONTENT_TYPE = "text/csv; charset=utf-8"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CATEGORY_SUFFIX = "_categoria"


@dataclass
class Table:
    headers: list[str]
    rows: list[dict[str, str]]


def parse_table(content: bytes, filename: str) -> Table:
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def serialize_table(table: Table, filename: str) -> tuple[bytes, str, str]:
    """Retorna (content, output_filename, content_type)."""
    if filename.lower().endswith(".xlsx"):
        content = _serialize_xlsx(table)
        return content, _result_filename(filename), XLSX_CONTENT_TYPE
    content = _serialize_csv(table)
    return content, _result_filename(filename), CSV_CONTENT_TYPE


def _parse_csv(content: bytes) -> Table:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append({h: ("" if row.get(h) is None else str(row[h])) for h in headers})
    return Table(headers=headers, rows=rows)


def _serialize_csv(table: Table) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=table.headers, lineterminator="\n")
    writer.writeheader()
    for row in table.rows:
        writer.writerow({h: row.get(h, "") for h in table.headers})
    return output.getvalue().encode("utf-8")


def _parse_xlsx(content: bytes) -> Table:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    raw_headers = [cell.value for cell in ws[1]]
    headers = [str(h) if h is not None else "" for h in raw_headers]

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2):
        row_dict: dict[str, str] = {}
        for idx, cell in enumerate(row):
            if idx >= len(headers):
                continue
            header = headers[idx]
            if not header:
                continue
            value = cell.value
            row_dict[header] = "" if value is None else str(value)
        rows.append(row_dict)
    wb.close()
    return Table(headers=headers, rows=rows)


def _serialize_xlsx(table: Table) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(table.headers)
    for row in table.rows:
        ws.append([row.get(h, "") for h in table.headers])
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


def _result_filename(original_filename: str) -> str:
    return f"normalized_{original_filename}"
