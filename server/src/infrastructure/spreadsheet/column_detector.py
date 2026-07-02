import csv
import io

from application.projects.detect_columns import ColumnDetector
from application.projects.dtos import DetectedColumnOutput


class SpreadsheetColumnDetector(ColumnDetector):
    _MAX_ROWS = 100
    _MAX_SAMPLES = 5

    def detect(self, content: bytes, filename: str) -> list[DetectedColumnOutput]:
        if filename.lower().endswith(".xlsx"):
            return self._detect_xlsx(content)
        return self._detect_csv(content)

    def _detect_csv(self, content: bytes) -> list[DetectedColumnOutput]:
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        columns: dict[str, list[str]] = {}
        for i, row in enumerate(reader):
            if i >= self._MAX_ROWS:
                break
            for col, val in row.items():
                if col not in columns:
                    columns[col] = []
                if val and len(columns[col]) < self._MAX_SAMPLES and val not in columns[col]:
                    columns[col].append(val)
        return [DetectedColumnOutput(column_name=k, sample_values=v) for k, v in columns.items()]

    def _detect_xlsx(self, content: bytes) -> list[DetectedColumnOutput]:
        from openpyxl import load_workbook  # deferred — only imported when needed

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(max_row=self._MAX_ROWS + 1, values_only=True))
        wb.close()

        if not rows:
            return []

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        columns: dict[str, list[str]] = {h: [] for h in headers}

        for row in rows[1:]:
            for h, val in zip(headers, row):
                if val is not None and len(columns[h]) < self._MAX_SAMPLES:
                    s = str(val)
                    if s not in columns[h]:
                        columns[h].append(s)

        return [DetectedColumnOutput(column_name=k, sample_values=v) for k, v in columns.items()]
