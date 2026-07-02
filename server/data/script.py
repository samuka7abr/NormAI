from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = Path(__file__).with_name("base_maus_tratos.xlsx")
DEFAULT_OUTPUT = Path(__file__).with_name("base_maus_tratos_5000.xlsx")
DEFAULT_DATA_ROWS = 5_000


def limit_xlsx_rows(input_path: Path, output_path: Path, data_rows: int) -> None:
    source = load_workbook(input_path, read_only=True, data_only=False)
    target = Workbook(write_only=True)

    try:
        for sheet_name in source.sheetnames:
            source_sheet = source[sheet_name]
            target_sheet = target.create_sheet(title=sheet_name)

            for index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                target_sheet.append(row)
                if index > data_rows:
                    break

        target.save(output_path)
    finally:
        source.close()
        target.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Cria uma copia XLSX mantendo o cabecalho e as primeiras N linhas de dados."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--rows", type=int, default=DEFAULT_DATA_ROWS)
    args = parser.parse_args()

    if args.rows < 1:
        raise SystemExit("--rows precisa ser maior que zero.")
    if not args.input.exists():
        raise SystemExit(f"Arquivo nao encontrado: {args.input}")

    limit_xlsx_rows(args.input, args.output, args.rows)
    print(f"Arquivo criado: {args.output}")


if __name__ == "__main__":
    main()
